import openpyxl

wb = openpyxl.load_workbook('Book.xlsx')

print(wb.sheetnames)

sheet = wb['Sheet1']
print(sheet['A1'])
print(sheet['A1'].value)


# Printing column A's first 3 rows
for i in range(1, 4):
    print(i, sheet.cell(row=i, column=1).value)


test_sheet = wb['TestSheet']

for i in range(1,4):
    print(i, test_sheet.cell(row=i, column=1).value)